"""Export inventory rows to Excel (.xlsx)."""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.inventory_service import inventory_stock_status


EXPORT_COLUMNS = (
    ('Product Name', lambda row: row['product_name']),
    ('SKU', lambda row: row['sku']),
    ('Size', lambda row: row['size']),
    ('Color', lambda row: row['color']),
    ('Category', lambda row: row['category']),
    ('Gender', lambda row: row['gender']),
    ('Age Groups', lambda row: row['age_groups']),
    ('Brand', lambda row: row['brand']),
    ('Price (INR)', lambda row: row['price']),
    ('Compare Price (INR)', lambda row: row['compare_at_price']),
    ('Cost Price (INR)', lambda row: row['cost_price']),
    ('Stock', lambda row: row['stock_quantity']),
    ('Status', lambda row: row['status']),
    ('Variant Active', lambda row: row['variant_active']),
    ('Product Slug', lambda row: row['slug']),
)


def variant_to_export_row(variant):
    """Convert a ProductVariant (with joined product) to a flat export dict."""
    product = variant.product
    return {
        'product_name': product.name,
        'sku': variant.sku,
        'size': variant.size,
        'color': variant.color,
        'category': product.category.name if product.category else '',
        'gender': product.gender_label or '',
        'age_groups': product.age_groups_label or '',
        'brand': product.brand or '',
        'price': float(product.price) if product.price is not None else None,
        'compare_at_price': float(product.compare_at_price) if product.compare_at_price else None,
        'cost_price': float(product.cost_price) if product.cost_price else None,
        'stock_quantity': variant.stock_quantity,
        'status': inventory_stock_status(variant.stock_quantity),
        'variant_active': 'Yes' if variant.is_active else 'No',
        'slug': product.slug,
    }


def export_inventory_xlsx(variants, filename_prefix='indistylex-inventory'):
    """
    Build an Excel workbook from inventory variants.

    Returns (BytesIO buffer, download_filename).
    """
    rows = [variant_to_export_row(variant) for variant in variants]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F2937')

    headers = [label for label, _ in EXPORT_COLUMNS]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    currency_cols = {9, 10, 11}
    for row in rows:
        values = [getter(row) for _, getter in EXPORT_COLUMNS]
        ws.append(values)
        current_row = ws.max_row
        for col_idx in currency_cols:
            cell = ws.cell(row=current_row, column=col_idx)
            if cell.value is not None:
                cell.number_format = '#,##0.00'
        ws.cell(row=current_row, column=12).number_format = '0'

    for col_idx, (header, _) in enumerate(EXPORT_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(header)
        for cell in ws[letter][1:]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.utcnow().strftime('%Y%m%d-%H%M%S')
    filename = f'{filename_prefix}-{timestamp}.xlsx'
    return buffer, filename

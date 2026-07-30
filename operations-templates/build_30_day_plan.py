#!/usr/bin/env python3
"""Generate Indistylex 30-Day Team Operations Plan Excel."""
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

START = date(2026, 7, 14)  # Day 1 — change if needed

# ── EDIT THESE NAMES & NUMBERS ──────────────────────────────────────
OWNER_NAME = 'Satyam Pandey'
OWNER_PHONE = '+91 [Your Number]'
MANAGER_NAME = '[Manager Name]'          # e.g. Rahul
MANAGER_PHONE = '+91 [Manager Number]'
ASSISTANT_NAME = '[Assistant Name]'      # e.g. Amit
ASSISTANT_PHONE = '+91 [Assistant Number]'
COMPANY_WHATSAPP = '+91 [Business Number]'  # Indistylex WhatsApp Business
# ───────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SUB_FILL = PatternFill('solid', fgColor='E5E7EB')
WRAP = Alignment(wrap_text=True, vertical='top')
THIN = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)


def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN


def auto_width(ws, max_width=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 12), max_width)


PLAN_ROWS = [
    # (week, owner, manager, assistant, excel, whatsapp_focus)
    (1, 'Create 2 WhatsApp groups. Share Excel on Google Drive.', 'Join groups. Read Team Roles tab. List all B2B shop contacts.', 'Shop stock count — every SKU. Photo of warehouse. Send count to group.', 'Inventory tab — enter all stock from shop count.', 'Group setup + first stock numbers'),
    (1, 'Verify admin login + B2B feature works. Save CREDENTIALS.', 'Login admin. Test Record B2B Sale (1 test sale).', 'Pack 3 sample pieces for shop display. Clean storage area.', 'Orders tab — add test B2B row. B2B Customers — add 3 shops.', 'Confirm admin works'),
    (1, 'Review website orders page. Check GA4 realtime once.', 'Morning: check admin orders 9 AM. Evening: 7 PM.', 'Visit 2 nearby shops — introduce Indistylex. Collect interest.', 'Expenses tab — add any shop costs (travel, packaging).', 'Manager posts order count'),
    (1, 'Instagram bio link → indistylex.com. Post 1 story.', 'Record shop visits in B2B Customers tab.', 'Deliver samples to 1 shop. Get shop owner WhatsApp.', 'B2B Customers — add shop phone + city + credit terms.', 'Assistant sends delivery proof photo'),
    (1, 'Weekly review call — 30 min (Sun). Fill Weekly KPIs Week 1.', 'Prepare Week 1 summary: orders, B2B leads, stock issues.', 'Full stock recount. Report low-stock SKUs.', 'Weekly KPIs tab — fill Week 1 row.', 'Week 1 summary in group'),
    (2, 'Approve B2B pricing rules (70% retail + extra discount limit).', 'Call 5 shops from list. Confirm orders for Week 2.', 'Deliver to 2 shops. Collect COD if any.', 'Orders tab — every B2B sale same day.', 'Daily update template — start habit'),
    (2, 'Fix any website issue from manager report.', 'Record all B2B in admin + Excel. Track credit vs COD.', 'Pack website orders if any. Hand to courier.', 'Inventory tab — reduce stock after each sale.', 'EOD update by 8 PM'),
    (2, 'Share product photos with manager for Instagram.', 'Post 1 feed post + 2 stories. Reply all DMs within 2h.', 'Visit 2 new shops. Note which SKUs they want.', 'Orders tab — note channel (B2B/Website).', 'Assistant: shops visited + qty'),
    (2, 'Check Razorpay + COD orders in admin.', 'Follow up unpaid B2B credit from last week.', 'Restock fast-moving sizes from warehouse.', 'Inventory — update Available to Sell column.', 'Manager: collections due list'),
    (2, 'Weekly review. Compare website vs B2B revenue.', 'Week 2 KPIs + top 5 SKUs report to owner.', 'Stock count on top 10 SKUs only (quick count).', 'Weekly KPIs Week 2.', 'Week 2 summary'),
    (3, 'Run Instagram Shopping setup (Commerce Manager).', 'Create 9-post grid plan. Schedule 3 posts.', 'Deliver to 3 shops. Target ₹10K+ B2B this week.', 'B2B Customers — mark active vs cold shops.', 'Mid-week revenue check'),
    (3, 'Review contact form + WhatsApp on website.', 'Process all website orders within 24h.', 'Photograph 5 products (phone camera, clean bg).', 'Expenses — shipping + packaging costs.', 'Photo dump to group'),
    (3, 'Google Business Profile — verify if not done.', 'Manager trains assistant: how to read SKU on tag.', 'Label all stock with SKU stickers if missing.', 'SKU & HSN tab — reference for invoices.', 'SKU question → manager answers'),
    (3, 'Check server + backup. Review admin inventory.', 'Low-stock alert: list SKUs under 5 units.', 'Priority delivery to shops that sell most.', 'Inventory Status column — fix Out of Stock.', 'Low stock alert in group'),
    (3, 'Weekly review. Decide Month 2 focus: B2B vs website.', 'Present: best shop, worst stock, revenue trend.', 'Shop feedback summary — what customers want.', 'Weekly KPIs Week 3 + Monthly P&L draft.', 'Week 3 summary'),
    (4, 'Plan Amazon listing — pick 5 hero SKUs.', 'Prepare Amazon Listing tab for 5 products.', 'Count stock reserved for website (do not oversell).', 'Amazon Listing tab — fill 5 rows.', 'Amazon prep status'),
    (4, 'Review GST invoices process for B2B.', 'Send payment reminders to credit shops.', 'Visit best 3 shops — reorder push.', 'B2B Customers — payment received column.', 'Collections update'),
    (4, 'Meta Pixel / ads — confirm or defer.', 'Instagram Reel #1 — packing or shop visit.', 'Pack all pending website orders.', 'Orders tab — catch up any missing rows.', 'Reel published?'),
    (4, 'Document what worked — 1 page notes for Month 2.', 'Manager writes SOP: daily routine (15 lines).', 'Assistant writes: delivery route + shop map.', 'All tabs reviewed for missing data.', 'SOP shared in group'),
    (4, '30-day review call — 1 hour. Set Month 2 targets.', 'Present full Month 1 report from Excel.', 'Stock audit — full count again.', 'Weekly KPIs Week 4 + Monthly P&L final.', 'Celebrate + Month 2 goals'),
]

# Expand to 30 days with rest days / lighter days
DAILY_PLAN = []
for day in range(1, 31):
    week = min((day - 1) // 7 + 1, 4)
    if day in (6, 13, 20, 27):
        row = PLAN_ROWS[(week - 1) * 5 + 4]
    else:
        within = (day - 1) % 7
        if within < 4:
            row = PLAN_ROWS[(week - 1) * 5 + within]
        else:
            row = PLAN_ROWS[(week - 1) * 5 + 3]
    DAILY_PLAN.append((day, week, row[1], row[2], row[3], row[4], row[5]))


def build_plan_sheet(wb):
    ws = wb.active
    ws.title = '30-Day Plan'
    headers = [
        'Day', 'Date', 'Week', 'Owner (Bengaluru)', 'Manager', 'Assistant (Allahabad Shop)',
        'Excel Tab Today', 'WhatsApp Update', 'Done? (Y/N)', 'Notes'
    ]
    ws.append(headers)
    style_header_row(ws)
    for day, week, owner, mgr, asst, excel, wa in DAILY_PLAN:
        d = START + timedelta(days=day - 1)
        ws.append([day, d.strftime('%d-%b-%Y'), f'Week {week}', owner, mgr, asst, excel, wa, '', ''])
    for row in ws.iter_rows(min_row=2, max_row=31):
        for cell in row:
            cell.alignment = WRAP
            cell.border = THIN
    ws.freeze_panes = 'A2'
    auto_width(ws)


def build_daily_checklist(wb):
    ws = wb.create_sheet('Daily Checklist')
    ws.append(['Role', 'Time', 'Task', 'Tool', 'Done?'])
    style_header_row(ws)
    rows = [
        ('Assistant (Allahabad)', '9:00 AM', 'Send opening message in WhatsApp group', 'WhatsApp', ''),
        ('Assistant (Allahabad)', '9:30 AM', 'Check pending deliveries from manager list', 'WhatsApp DM', ''),
        ('Assistant (Allahabad)', '10 AM–6 PM', 'Deliveries, shop visits, packing, stock count', 'Physical', ''),
        ('Assistant (Allahabad)', '6:30 PM', 'Send EOD update (template in WhatsApp Templates tab)', 'WhatsApp', ''),
        ('Assistant (Allahabad)', '7:00 PM', 'Hand cash collections to manager (bank transfer same day)', 'PhonePe/Bank', ''),
        ('Manager', '9:00 AM', 'Check admin → Orders (new website orders)', 'indistylex.com/admin', ''),
        ('Manager', '9:15 AM', 'Check admin → Inventory (low stock)', 'indistylex.com/admin', ''),
        ('Manager', '10:00 AM', 'Send today task list to assistant on WhatsApp', 'WhatsApp', ''),
        ('Manager', '12:00 PM', 'Record any B2B sale in admin + Excel Orders tab', 'Admin + Excel', ''),
        ('Manager', '4:00 PM', 'Follow up shops — credit payments, new orders', 'Phone/WhatsApp', ''),
        ('Manager', '7:30 PM', 'Update Excel (Orders, Inventory, Expenses)', 'Excel', ''),
        ('Manager', '8:00 PM', 'Post team summary in WhatsApp group', 'WhatsApp', ''),
        ('Owner (Bengaluru)', '9:00 AM', 'Read WhatsApp group — reply blockers only', 'WhatsApp', ''),
        ('Owner (Bengaluru)', 'Mon 10 AM', 'Weekly KPI review — 30 min call with manager', 'Excel + Call', ''),
        ('Owner (Bengaluru)', 'As needed', 'Website/server fixes, approve big discounts', 'Admin/Server', ''),
        ('Owner (Bengaluru)', 'Sun 7 PM', 'Weekly review — fill Weekly KPIs tab', 'Excel', ''),
    ]
    for r in rows:
        ws.append(list(r))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = THIN
    auto_width(ws)


def build_weekly_review(wb):
    ws = wb.create_sheet('Weekly Review')
    ws.append(['Week', 'Dates', 'Website Orders', 'B2B Orders', 'B2B Revenue ₹', 'Website Revenue ₹',
               'Total Expenses ₹', 'Net Profit ₹', 'Shops Visited', 'New B2B Shops', 'Top SKU',
               'Biggest Problem', 'Next Week Priority', 'Owner Sign-off'])
    style_header_row(ws)
    for w in range(1, 5):
        start = START + timedelta(days=(w - 1) * 7)
        end = start + timedelta(days=6)
        ws.append([f'Week {w}', f'{start.strftime("%d %b")} – {end.strftime("%d %b")}',
                   '', '', '', '', '', '', '', '', '', '', '', ''])
    auto_width(ws)


def build_team_contacts(wb):
    ws = wb.create_sheet('Team Contacts', 0)
    ws.append(['Field', 'Value', 'Notes'])
    style_header_row(ws)
    rows = [
        ('Owner', OWNER_NAME, 'Bengaluru — tech, strategy, weekly review'),
        ('Owner WhatsApp', OWNER_PHONE, ''),
        ('Manager', MANAGER_NAME, 'Admin, Excel, B2B, Instagram, daily updates'),
        ('Manager WhatsApp', MANAGER_PHONE, ''),
        ('Assistant (Shop)', ASSISTANT_NAME, 'Allahabad — deliveries, stock, shop visits'),
        ('Assistant WhatsApp', ASSISTANT_PHONE, ''),
        ('Business WhatsApp', COMPANY_WHATSAPP, 'Customer-facing Indistylex number'),
        ('Admin URL', 'https://indistylex.com/admin', 'Manager login only'),
        ('Excel on Drive', '[Paste Google Drive link]', 'Share with Manager (Editor)'),
        ('Weekly call', 'Sunday 7:00 PM IST', 'Owner + Manager, 30 min'),
    ]
    for r in rows:
        ws.append(list(r))
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = THIN


def build_whatsapp_groups(wb):
    ws = wb.create_sheet('WhatsApp Group Setup')
    ws.append(['Group Name', 'Members', 'Group Description (copy-paste)', 'Group Rules (pin as message)'])
    style_header_row(ws)

    g1_desc = (
        f'Indistylex core team — daily operations\n'
        f'👤 Owner: {OWNER_NAME} (Bengaluru)\n'
        f'👤 Manager: {MANAGER_NAME}\n'
        f'👤 Shop: {ASSISTANT_NAME} (Allahabad)\n\n'
        f'📋 Daily updates required\n'
        f'🕕 Assistant EOD: 6:30 PM\n'
        f'🕗 Manager summary: 8:00 PM\n'
        f'📗 Excel + admin sync same day'
    )
    g1_rules = (
        'GROUP RULES — Indistylex Team\n\n'
        '1️⃣ Assistant posts EOD update by 6:30 PM\n'
        '2️⃣ Manager posts daily summary by 8:00 PM\n'
        '3️⃣ Every delivery = photo in group\n'
        '4️⃣ Stock issues → tag @Manager\n'
        '5️⃣ Tech/payment issues → tag @Owner\n'
        '6️⃣ Sunday = weekly summary from Manager\n\n'
        'Templates: see Excel → WhatsApp Templates tab'
    )

    g2_desc = (
        f'Fast shop coordination — Manager + Assistant only\n'
        f'Manager: {MANAGER_NAME}\n'
        f'Assistant: {ASSISTANT_NAME} (Allahabad shop)\n\n'
        f'Use for: delivery tasks, packing, urgent stock, photos'
    )
    g2_rules = (
        'SHOP OPS RULES\n\n'
        '1️⃣ Manager sends task list every morning by 10 AM\n'
        '2️⃣ Assistant replies ✅ when task read\n'
        '3️⃣ Photo proof for every delivery\n'
        '4️⃣ Cash collected → report amount + transfer same day'
    )

    g3_desc = (
        'B2B wholesale shop owners — Manager handles only\n'
        'Order confirmations, payment reminders, new arrivals\n'
        'Website: indistylex.com'
    )
    g3_rules = (
        'Welcome to Indistylex wholesale!\n\n'
        '🏷️ Kids clothing 0-14 years\n'
        '🚚 Delivery in Prayagraj / nearby\n'
        '💳 COD & credit available (approved shops)\n'
        '📦 Min order: [set your minimum]\n'
        '👤 Your manager contact: ' + MANAGER_NAME
    )

    data = [
        ('Indistylex — Team', f'{OWNER_NAME}, {MANAGER_NAME}, {ASSISTANT_NAME}', g1_desc, g1_rules),
        ('Indistylex — Shop Ops', f'{MANAGER_NAME}, {ASSISTANT_NAME}', g2_desc, g2_rules),
        ('Indistylex — B2B Shops', f'{MANAGER_NAME} + shop owners', g3_desc, g3_rules),
    ]
    for d in data:
        ws.append(list(d))
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 45
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = THIN


def build_team_roles(wb):
    ws = wb.create_sheet('Team Roles')
    ws.append(['Role', 'Location', 'Person', 'Phone', 'Responsibilities', 'Tools', 'Does NOT do'])
    style_header_row(ws)
    data = [
        ('Owner', 'Bengaluru', OWNER_NAME, OWNER_PHONE,
         'Strategy, website/server, payments, final approvals, weekly review',
         'Admin, server, Excel review, WhatsApp',
         'Daily shop visits, daily Excel entry'),
        ('Manager', 'Remote / Bengaluru', MANAGER_NAME, MANAGER_PHONE,
         'Admin panel, Excel updates, B2B calls, shop coordination, Instagram posts, daily group updates',
         'Admin, Excel, WhatsApp, Instagram',
         'Server/code, heavy lifting'),
        ('Assistant', 'Allahabad (Shop)', ASSISTANT_NAME, ASSISTANT_PHONE,
         'Deliveries, shop visits, packing, stock count, COD collection, photos',
         'WhatsApp, phone, physical stock',
         'Admin login, pricing decisions, Excel (unless trained)'),
    ]
    for r in data:
        ws.append(list(r))
    auto_width(ws)


def build_whatsapp(wb):
    ws = wb.create_sheet('WhatsApp Templates')
    ws.append(['Template Name', 'Who Posts', 'When', 'Message (copy-paste)'])
    style_header_row(ws)
    templates = [
        ('Morning — Assistant', 'Assistant', '9:00 AM',
         '🌅 Good morning team!\n📍 Allahabad Shop\n✅ Ready for today\n📦 Deliveries planned: [list]\n❓ Issues: None'),
        ('EOD — Assistant', 'Assistant', '6:30 PM',
         '📋 EOD Update — [DATE]\n🏪 Shops visited: [names]\n📦 Delivered: [qty] pcs to [shop]\n💰 COD collected: ₹[amount]\n📊 Stock issues: [low SKU list]\n📸 Photos: shared'),
        ('EOD — Manager', 'Manager', '8:00 PM',
         '📊 Daily Summary — [DATE]\n🌐 Website orders: [count] — ₹[amount]\n🏪 B2B sales: [count] — ₹[amount]\n📝 Admin updated: ✅\n📗 Excel updated: ✅\n⚠️ Tomorrow: [tasks for assistant]'),
        ('Weekly — Manager', 'Manager', 'Sunday',
         '📈 Week [N] Summary\n🌐 Website: [orders] orders — ₹[rev]\n🏪 B2B: [orders] — ₹[rev]\n🏆 Best shop: [name]\n⚠️ Problem: [issue]\n🎯 Next week: [priority]'),
        ('Urgent — Stockout', 'Anyone', 'Anytime',
         '🚨 LOW STOCK: [SKU] — only [X] left\n@Manager please update website/admin'),
        ('Task — Manager to Assistant', 'Manager', 'Morning',
         '📌 Today\'s tasks:\n1. Deliver [qty] to [shop name] — [address]\n2. Pack website order #[order]\n3. Count stock for [SKU list]\nReply ✅ when read'),
    ]
    for t in templates:
        ws.append(list(t))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = THIN
    ws.column_dimensions['D'].width = 60
    auto_width(ws, 60)


def build_excel_guide(wb):
    ws = wb.create_sheet('How to Use Tracker')
    ws.append(['Tab in Business-Tracker.xlsx', 'Who Updates', 'When', 'What to Enter'])
    style_header_row(ws)
    guide = [
        ('Inventory', 'Manager', 'Daily + after every sale', 'Stock qty per SKU. Match admin after B2B/website sale.'),
        ('Orders', 'Manager', 'Same day as sale', 'One row per order. Channel: Website/B2B/Amazon/Flipkart.'),
        ('Expenses', 'Manager', 'When money spent', 'Wholesaler, courier, packaging, ads, travel.'),
        ('Weekly KPIs', 'Owner + Manager', 'Every Sunday', 'Copy totals from Orders. Compare to targets.'),
        ('B2B Customers', 'Manager', 'Per shop contact', 'Shop name, phone, city, credit terms, payment status.'),
        ('SKU & HSN Codes', 'Reference', 'As needed', 'Look up HSN for invoices. Do not edit unless new product.'),
        ('Monthly P&L', 'Owner', '1st of month', 'Revenue − expenses by channel.'),
        ('Amazon / Flipkart Listing', 'Owner', 'When listing', 'Product data for marketplace upload.'),
    ]
    for g in guide:
        ws.append(list(g))
    auto_width(ws)


def build_communication(wb):
    ws = wb.create_sheet('Communication Rules')
    ws.append(['Rule', 'Detail'])
    style_header_row(ws)
    rules = [
        ('Group 1: Indistylex — Team', 'Owner + Manager + Assistant. Daily updates. Owner reads, Manager leads.'),
        ('Group 2: Indistylex — Shop Ops', 'Manager + Assistant only. Fast task coordination.'),
        ('Group 3: Indistylex — B2B Shops', 'Manager only adds shop owners. No assistant personal number exposed.'),
        ('Daily update deadline', 'Assistant 6:30 PM IST | Manager 8:00 PM IST'),
        ('Weekly call', 'Sunday 7 PM — Owner + Manager (30 min). Assistant sends written summary before call.'),
        ('Excel is source of truth for money', 'Admin = stock truth. Excel = accounting/planning. Sync same day.'),
        ('Escalation to Owner', 'Server down, payment issue, discount >10%, shop credit dispute'),
        ('Photos required', 'Every delivery — photo in group. Every stock count — photo of sheet.'),
        ('Cash handling', 'Assistant collects → transfers to company account same day → Manager records in Excel'),
        ('Missed update?', 'Manager calls assistant at 7 PM. No update 2 days = owner joins call.'),
    ]
    for r in rules:
        ws.append(list(r))
    auto_width(ws, 70)


def main():
    wb = Workbook()
    build_team_contacts(wb)
    build_plan_sheet(wb)
    build_daily_checklist(wb)
    build_weekly_review(wb)
    build_whatsapp_groups(wb)
    build_team_roles(wb)
    build_whatsapp(wb)
    build_excel_guide(wb)
    build_communication(wb)
    out = 'Indistylex-30-Day-Team-Plan.xlsx'
    wb.save(out)
    print(f'Created {out}')
    print(f'  Team: {OWNER_NAME} | {MANAGER_NAME} | {ASSISTANT_NAME}')
    print('  Edit names at top of build_30_day_plan.py then re-run to regenerate.')


if __name__ == '__main__':
    main()
